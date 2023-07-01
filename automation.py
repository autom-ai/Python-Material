# -*- coding: utf-8 -*-
"""
Created on Sat Apr 29 14:24:25 2023

@author: Mini Einstein
"""

import openpyxl as xl
wb=xl.load_workbook('transactions.xlsx') 
st=wb['Sheet1']
for i in range(2,st.max_row + 1):
    cell1=st.cell(i,3)
    correct=cell1.value * 0.9
    correct_cell=st.cell(i,4)
    correct_cell.value=correct
wb.save('transactions.xlsx')
